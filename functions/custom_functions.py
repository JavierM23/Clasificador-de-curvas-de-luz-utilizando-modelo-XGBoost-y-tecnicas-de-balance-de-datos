import numpy as np
import xgboost as xgb
import pickle
from openpyxl import load_workbook
from metrics import IAM, CBA, recall, MPAUC, MAvG, f1, kappa, MPR, MCC, min_recall
import json
from sklearn.neighbors import NearestNeighbors
import pandas as pd
from functools import partial
from focal_loss_CE import focal_loss_obj
from collections import Counter


def pickle_save(file_name: str, pickle_dict):
    with open(file_name, 'wb') as pickle_file:
        model_dump = pickle_dict
        pickle.dump(model_dump, pickle_file)
    return None

def pickle_load(file_name:str):
    pickle_dict = pickle.load(open(file_name, "rb" ))
    return pickle_dict

def xgb_train(params, dtrain, dtest, feval='recall', verbose_eval=True, nthread=-1, FL = False, gamma = 0):

    feval_dict = {'IAM': IAM, 'CBA': CBA, 'recall': recall, 'MPAUC': MPAUC, 'MAvG':MAvG, 'f1':f1, 'kappa':kappa, 'MPR':MPR, 'MCC':MCC, 'min_recall':min_recall}

    es = xgb.callback.EarlyStopping(rounds=50, save_best=True, maximize=False, data_name="test", metric_name=feval)
    nclass = len(np.unique(np.array(dtrain.get_label(), dtype=int, ndmin=1)))

    model_params = {
        'max_depth':int(params['max_depth']),
        'min_child_weight':int(params['min_child_weight']),
        'learning_rate':params['learning_rate'],
        'objective':'multi:softprob',
        'gamma':params['gamma'],
        'max_delta_step': 0,
        'subsample':params['subsample'],
        'colsample_bytree':params['colsample_bytree'],
        #'colsample_bynode':params['colsample_bynode'],#
        #'colsample_bylevel':params['colsample_bylevel'],#
        'reg_alpha':params['reg_alpha'],
        'reg_lambda':params['reg_lambda'],
        'scale_pos_weight':1,
        'missing':-999,
        'n_jobs':5,
        'num_class':nclass,
        'nthread':nthread}
        
    if FL:
        focal_loss = partial(focal_loss_obj, gamma=gamma)
        xgb_model = xgb.train(model_params, dtrain, num_boost_round=300, obj=focal_loss,
                        evals=[(dtest, 'test')], feval=feval_dict[feval], callbacks=[es], verbose_eval=verbose_eval)
    else:
        xgb_model = xgb.train(model_params, dtrain, num_boost_round=300,
                        evals=[(dtest, 'test')], feval=feval_dict[feval], callbacks=[es], verbose_eval=verbose_eval)
    return xgb_model

def prediction_union(y_pred_hierarchical, y_pred_periodic, y_pred_stochastic, y_pred_transient):
    prob_periodic = y_pred_periodic*np.array(y_pred_hierarchical[:,2], ndmin=2).T
    prob_stochastic = y_pred_stochastic*np.array(y_pred_hierarchical[:,1], ndmin=2).T
    prob_transient = y_pred_transient*np.array(y_pred_hierarchical[:,0], ndmin=2).T
    prob_final = np.concatenate((prob_transient, prob_stochastic, prob_periodic) ,axis=1)
    #class_final_index = np.argmax(prob_final,axis=1)
    return prob_final

def save_performance(model, method, subclass, metric_name, Metrics, cnf_metrics):
    new_row = [model, method, subclass, metric_name,
    Metrics[0], Metrics[1], Metrics[2], Metrics[3], Metrics[4], Metrics[5], Metrics[6], Metrics[7],
    cnf_metrics[0], cnf_metrics[1], cnf_metrics[2], cnf_metrics[3], cnf_metrics[4]]
    wb = load_workbook("model_metrics.xlsx")
    ws = wb.worksheets[0]
    ws.append(new_row)
    wb.save("model_metrics.xlsx")
    return None

def save_excel(excel_file, sheet=0, *args):
    new_row = []
    for x in args:
        if type(x)==list:
            new_row += x
        else:
            new_row += [x]
    wb = load_workbook(excel_file)
    ws = wb.worksheets[sheet]
    ws.append(new_row)
    wb.save(excel_file)
    return None


def get_strategy(y, rep_class, kind: str = 'over'):
    count = Counter(y)
    if kind == 'over':
        strategy = {i:count[i] for i in range(rep_class+1)}
        for i in range(rep_class+1,len(count)):
            strategy[i]=count[rep_class]
    elif kind == 'under':
        strategy = {i:count[rep_class] for i in range(rep_class+1)}
        for i in range(rep_class+1,len(count)):
            strategy[i]=count[i]
    else:
        raise Warning("kind = {'over', 'under'}")
    return strategy

